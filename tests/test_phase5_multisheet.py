"""Phase 5 tests: Multi-sheet support, XLSX import, font color formatting."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest
import yaml

from fin123.project import scaffold_project
from fin123.ui.service import ProjectService


# ────────────────────────────────────────────────────────────────
# Fixtures
# ────────────────────────────────────────────────────────────────


@pytest.fixture
def demo_project(tmp_path: Path) -> Path:
    """Scaffold a demo project for testing."""
    project_dir = tmp_path / "test_project"
    scaffold_project(project_dir)
    return project_dir


@pytest.fixture
def service(demo_project: Path) -> ProjectService:
    """Create a ProjectService for testing."""
    return ProjectService(project_dir=demo_project)


@pytest.fixture
def multi_sheet_project(tmp_path: Path) -> Path:
    """Create a project with multiple sheets pre-configured."""
    d = tmp_path / "multi"
    d.mkdir()
    spec = {
        "sheets": [
            {"name": "Revenue", "n_rows": 200, "n_cols": 40, "cells": {"A1": {"value": 100}}},
            {"name": "Costs", "n_rows": 200, "n_cols": 40, "cells": {"A1": {"value": 50}}},
        ],
        "params": {},
        "tables": {},
        "plans": [],
        "outputs": [],
    }
    (d / "workbook.yaml").write_text(yaml.dump(spec, default_flow_style=False))
    (d / "fin123.yaml").write_text(yaml.dump({"max_runs": 50}))
    for sub in ["inputs", "runs", "artifacts", "snapshots", "cache"]:
        (d / sub).mkdir()
    return d


# ────────────────────────────────────────────────────────────────
# Multi-sheet CRUD
# ────────────────────────────────────────────────────────────────


class TestMultiSheetCRUD:
    def test_list_sheets_default(self, service: ProjectService) -> None:
        """Demo project defaults to one Sheet1."""
        sheets = service.list_sheets()
        assert len(sheets) >= 1
        assert sheets[0]["name"] == "Sheet1"

    def test_add_sheet(self, service: ProjectService) -> None:
        result = service.add_sheet("Sheet2")
        assert result["name"] == "Sheet2"
        sheets = service.list_sheets()
        names = [s["name"] for s in sheets]
        assert "Sheet2" in names

    def test_add_duplicate_sheet(self, service: ProjectService) -> None:
        with pytest.raises(ValueError, match="already exists"):
            service.add_sheet("Sheet1")

    def test_delete_sheet(self, service: ProjectService) -> None:
        service.add_sheet("Sheet2")
        result = service.delete_sheet("Sheet2")
        assert result["deleted"] == "Sheet2"
        assert "Sheet2" not in result["remaining"]

    def test_delete_only_sheet(self, service: ProjectService) -> None:
        with pytest.raises(ValueError, match="Cannot delete"):
            service.delete_sheet("Sheet1")

    def test_delete_nonexistent_sheet(self, service: ProjectService) -> None:
        service.add_sheet("Sheet2")  # Need 2+ sheets so "only sheet" check doesn't fire first
        with pytest.raises(ValueError, match="not found"):
            service.delete_sheet("NoSuch")

    def test_rename_sheet(self, service: ProjectService) -> None:
        result = service.rename_sheet("Sheet1", "Main")
        assert result["new_name"] == "Main"
        sheets = service.list_sheets()
        assert sheets[0]["name"] == "Main"

    def test_rename_to_existing(self, service: ProjectService) -> None:
        service.add_sheet("Sheet2")
        with pytest.raises(ValueError, match="already exists"):
            service.rename_sheet("Sheet1", "Sheet2")

    def test_project_info_lists_sheets(self, service: ProjectService) -> None:
        service.add_sheet("Sheet2")
        info = service.get_project_info()
        assert "Sheet1" in info["sheets"]
        assert "Sheet2" in info["sheets"]


# ────────────────────────────────────────────────────────────────
# Independent per-sheet editing
# ────────────────────────────────────────────────────────────────


class TestPerSheetEditing:
    def test_edit_different_sheets_independently(self, service: ProjectService) -> None:
        service.add_sheet("Sheet2")

        # Edit Sheet1 cell A1
        service.update_cells("Sheet1", [{"addr": "A1", "value": "hello"}])
        # Edit Sheet2 cell A1
        service.update_cells("Sheet2", [{"addr": "A1", "value": "world"}])

        vp1 = service.get_sheet_viewport("Sheet1", 0, 0, 10, 10)
        vp2 = service.get_sheet_viewport("Sheet2", 0, 0, 10, 10)

        cells1 = {c["addr"]: c for c in vp1["cells"]}
        cells2 = {c["addr"]: c for c in vp2["cells"]}

        assert cells1["A1"]["display"] == "hello"
        assert cells2["A1"]["display"] == "world"

    def test_multi_sheet_project_loads(self, multi_sheet_project: Path) -> None:
        svc = ProjectService(project_dir=multi_sheet_project)
        sheets = svc.list_sheets()
        assert len(sheets) == 2
        assert sheets[0]["name"] == "Revenue"
        assert sheets[1]["name"] == "Costs"

    def test_multi_sheet_viewport(self, multi_sheet_project: Path) -> None:
        svc = ProjectService(project_dir=multi_sheet_project)
        vp = svc.get_sheet_viewport("Revenue", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        assert cells["A1"]["display"] == "100"


# ────────────────────────────────────────────────────────────────
# Snapshot persistence
# ────────────────────────────────────────────────────────────────


class TestSnapshotPersistence:
    def test_save_preserves_multi_sheet(self, service: ProjectService) -> None:
        service.add_sheet("Sheet2")
        service.update_cells("Sheet2", [{"addr": "B2", "value": "42"}])
        result = service.save_snapshot()
        assert result["snapshot_version"]

        # Reload from disk
        svc2 = ProjectService(project_dir=service.project_dir)
        sheets = svc2.list_sheets()
        names = [s["name"] for s in sheets]
        assert "Sheet2" in names

        vp = svc2.get_sheet_viewport("Sheet2", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        assert cells["B2"]["display"] == "42"

    def test_save_preserves_fmt(self, service: ProjectService) -> None:
        service.update_cells("Sheet1", [{"addr": "A1", "value": "100"}])
        service.update_cell_format("Sheet1", [{"addr": "A1", "color": "#ff0000"}])
        service.save_snapshot()

        svc2 = ProjectService(project_dir=service.project_dir)
        vp = svc2.get_sheet_viewport("Sheet1", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        assert cells["A1"]["fmt"]["color"] == "#ff0000"


# ────────────────────────────────────────────────────────────────
# Backward compatibility
# ────────────────────────────────────────────────────────────────


class TestBackwardCompat:
    def test_no_sheets_key_defaults_to_sheet1(self, tmp_path: Path) -> None:
        """A workbook.yaml without 'sheets' key should still work."""
        d = tmp_path / "legacy"
        d.mkdir()
        spec = {"params": {}, "tables": {}, "plans": [], "outputs": []}
        (d / "workbook.yaml").write_text(yaml.dump(spec))
        (d / "fin123.yaml").write_text(yaml.dump({"max_runs": 50}))
        for sub in ["inputs", "runs", "artifacts", "snapshots", "cache"]:
            (d / sub).mkdir()

        svc = ProjectService(project_dir=d)
        sheets = svc.list_sheets()
        assert len(sheets) == 1
        assert sheets[0]["name"] == "Sheet1"

    def test_demo_project_still_works(self, demo_project: Path) -> None:
        """Demo project with existing format should load correctly."""
        svc = ProjectService(project_dir=demo_project)
        info = svc.get_project_info()
        assert info["engine_version"]
        vp = svc.get_sheet_viewport("Sheet1", 0, 0, 10, 10)
        assert vp["n_rows"] > 0


# ────────────────────────────────────────────────────────────────
# Font color formatting
# ────────────────────────────────────────────────────────────────


class TestFontColor:
    def test_set_color(self, service: ProjectService) -> None:
        service.update_cells("Sheet1", [{"addr": "A1", "value": "100"}])
        result = service.update_cell_format("Sheet1", [{"addr": "A1", "color": "#4f7cff"}])
        assert result["ok"]
        assert result["dirty"]

    def test_color_in_viewport(self, service: ProjectService) -> None:
        service.update_cells("Sheet1", [{"addr": "A1", "value": "100"}])
        service.update_cell_format("Sheet1", [{"addr": "A1", "color": "#ff5c5c"}])
        vp = service.get_sheet_viewport("Sheet1", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        assert cells["A1"]["fmt"]["color"] == "#ff5c5c"

    def test_clear_color(self, service: ProjectService) -> None:
        service.update_cells("Sheet1", [{"addr": "A1", "value": "100"}])
        service.update_cell_format("Sheet1", [{"addr": "A1", "color": "#ff0000"}])
        service.update_cell_format("Sheet1", [{"addr": "A1", "color": None}])
        vp = service.get_sheet_viewport("Sheet1", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        # Cell exists but no fmt
        assert "fmt" not in cells.get("A1", {}) or cells["A1"].get("fmt") is None

    def test_color_only_cell_in_viewport(self, service: ProjectService) -> None:
        """A cell with only fmt (no value) should appear in viewport."""
        service.update_cell_format("Sheet1", [{"addr": "B2", "color": "#00ff00"}])
        vp = service.get_sheet_viewport("Sheet1", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        assert "B2" in cells
        assert cells["B2"]["fmt"]["color"] == "#00ff00"

    def test_color_persists_after_save(self, service: ProjectService) -> None:
        service.update_cells("Sheet1", [{"addr": "C3", "value": "test"}])
        service.update_cell_format("Sheet1", [{"addr": "C3", "color": "#ff5c5c"}])
        service.save_snapshot()

        svc2 = ProjectService(project_dir=service.project_dir)
        vp = svc2.get_sheet_viewport("Sheet1", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        assert cells["C3"]["fmt"]["color"] == "#ff5c5c"


# ────────────────────────────────────────────────────────────────
# XLSX Import
# ────────────────────────────────────────────────────────────────


class TestXlsxImport:
    def _create_test_xlsx(self, path: Path) -> None:
        """Create a minimal test .xlsx file using openpyxl."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1["A1"] = 100
        ws1["A2"] = "hello"
        ws1["B1"] = 200.5
        ws1["C1"] = "=A1+B1"

        # Add font color
        from openpyxl.styles import Font
        ws1["A1"].font = Font(color="FF0000")

        # Second sheet
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Total"
        ws2["B1"] = "=Data!A1"  # Cross-sheet ref (will be kept as-is)

        wb.save(str(path))

    def test_basic_import(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "test.xlsx"
        self._create_test_xlsx(xlsx)

        target = tmp_path / "imported"
        from fin123.xlsx_import import import_xlsx

        report = import_xlsx(xlsx, target)

        assert len(report["sheets_imported"]) == 2
        assert report["cells_imported"] > 0
        assert (target / "workbook.yaml").exists()
        assert (target / "import_report.json").exists()

    def test_import_creates_valid_project(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "test.xlsx"
        self._create_test_xlsx(xlsx)

        target = tmp_path / "imported"
        from fin123.xlsx_import import import_xlsx

        import_xlsx(xlsx, target)

        # Should be loadable as ProjectService
        svc = ProjectService(project_dir=target)
        sheets = svc.list_sheets()
        names = [s["name"] for s in sheets]
        assert "Data" in names
        assert "Summary" in names

    def test_import_values(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "test.xlsx"
        self._create_test_xlsx(xlsx)

        target = tmp_path / "imported"
        from fin123.xlsx_import import import_xlsx

        import_xlsx(xlsx, target)

        svc = ProjectService(project_dir=target)
        vp = svc.get_sheet_viewport("Data", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        assert cells["A1"]["display"] == "100"
        assert cells["A2"]["display"] == "hello"

    def test_import_formula(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "test.xlsx"
        self._create_test_xlsx(xlsx)

        target = tmp_path / "imported"
        from fin123.xlsx_import import import_xlsx

        import_xlsx(xlsx, target)

        svc = ProjectService(project_dir=target)
        vp = svc.get_sheet_viewport("Data", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        # Formula should be preserved
        assert cells["C1"]["raw"].startswith("=")

    def test_import_color(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "test.xlsx"
        self._create_test_xlsx(xlsx)

        target = tmp_path / "imported"
        from fin123.xlsx_import import import_xlsx

        import_xlsx(xlsx, target)

        svc = ProjectService(project_dir=target)
        vp = svc.get_sheet_viewport("Data", 0, 0, 10, 10)
        cells = {c["addr"]: c for c in vp["cells"]}
        assert "fmt" in cells["A1"]
        assert cells["A1"]["fmt"]["color"].startswith("#")

    def test_import_report_json(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "test.xlsx"
        self._create_test_xlsx(xlsx)

        target = tmp_path / "imported"
        from fin123.xlsx_import import import_xlsx

        import_xlsx(xlsx, target)

        report = json.loads((target / "import_report.json").read_text())
        assert "sheets_imported" in report
        assert report["cells_imported"] > 0

    def test_import_max_rows(self, tmp_path: Path) -> None:
        """Verify max_rows limit is respected."""
        import openpyxl

        xlsx = tmp_path / "big.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 20):
            ws.cell(row=i, column=1, value=i)
        wb.save(str(xlsx))

        target = tmp_path / "limited"
        from fin123.xlsx_import import import_xlsx

        report = import_xlsx(xlsx, target, max_rows=5)
        assert report["sheets_imported"][0]["cells"] == 5


# ────────────────────────────────────────────────────────────────
# API endpoints (via httpx TestClient)
# ────────────────────────────────────────────────────────────────


class TestAPIEndpoints:
    @pytest.fixture
    def client(self, demo_project: Path):
        from fastapi.testclient import TestClient
        from fin123.ui.server import create_app

        app = create_app(demo_project)
        return TestClient(app)

    def test_list_sheets(self, client) -> None:
        resp = client.get("/api/sheets")
        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, list)
        assert data[0]["name"] == "Sheet1"

    def test_add_sheet(self, client) -> None:
        resp = client.post("/api/sheets", json={"name": "Sheet2"})
        assert resp.status_code == 200
        assert resp.json()["name"] == "Sheet2"

    def test_add_duplicate_sheet(self, client) -> None:
        resp = client.post("/api/sheets", json={"name": "Sheet1"})
        assert resp.status_code == 400

    def test_delete_sheet(self, client) -> None:
        client.post("/api/sheets", json={"name": "Sheet2"})
        resp = client.request("DELETE", "/api/sheets", json={"name": "Sheet2"})
        assert resp.status_code == 200

    def test_rename_sheet(self, client) -> None:
        resp = client.patch("/api/sheets", json={"old_name": "Sheet1", "new_name": "Main"})
        assert resp.status_code == 200
        assert resp.json()["new_name"] == "Main"

    def test_format_endpoint(self, client) -> None:
        # Set a value first
        client.post("/api/sheet/cells", json={
            "sheet": "Sheet1",
            "edits": [{"addr": "A1", "value": "100"}]
        })
        # Set color
        resp = client.post("/api/sheet/format", json={
            "sheet": "Sheet1",
            "updates": [{"addr": "A1", "color": "#ff0000"}]
        })
        assert resp.status_code == 200
        assert resp.json()["ok"]

        # Verify in viewport
        resp = client.get("/api/sheet?sheet=Sheet1&r0=0&c0=0&rows=10&cols=10")
        data = resp.json()
        cells = {c["addr"]: c for c in data["cells"]}
        assert cells["A1"]["fmt"]["color"] == "#ff0000"

    def test_multi_sheet_viewport(self, client) -> None:
        # Add sheet
        client.post("/api/sheets", json={"name": "Sheet2"})
        # Edit on Sheet2
        client.post("/api/sheet/cells", json={
            "sheet": "Sheet2",
            "edits": [{"addr": "A1", "value": "999"}]
        })
        # Verify
        resp = client.get("/api/sheet?sheet=Sheet2&r0=0&c0=0&rows=10&cols=10")
        data = resp.json()
        cells = {c["addr"]: c for c in data["cells"]}
        assert cells["A1"]["display"] == "999"
