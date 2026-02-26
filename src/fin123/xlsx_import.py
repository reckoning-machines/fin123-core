"""XLSX import — best-effort conversion of Excel workbooks to fin123 format.

Uses openpyxl to read .xlsx files and produces a workbook.yaml-compatible
sheet structure.  Extracts:
- Cell values (numbers, strings, booleans)
- Formulas (prefixed with ``=``, as-is — no translation)
- Font color (stored as ``fmt.color`` hex)

Produces an ``import_report.json`` summarising what was imported and what
was skipped (charts, pivot tables, VBA, conditional formatting, etc.).
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from lark import Visitor, Tree

from fin123.formulas.parser import parse_formula
from fin123.formulas.evaluator import _FUNC_TABLE


def _col_letter(idx: int) -> str:
    """Convert 0-based column index to letter(s)."""
    result = ""
    n = idx + 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _make_addr(row: int, col: int) -> str:
    """Build cell address from 0-based row/col."""
    return f"{_col_letter(col)}{row + 1}"


_EXCEL_REF_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+")

_SUPPORTED_FUNCTIONS = set(_FUNC_TABLE.keys())

_EXTERNAL_LINK_RE = re.compile(
    r"\[.*?\]"        # [Workbook.xlsx]Sheet!A1
    r"|https?://"     # http:// or https:// URLs
    r"|'[A-Z]:\\"     # 'C:\path\...'
    r"|\\\\\\"        # \\server\path UNC
)

_PLUGIN_PREFIXES = ("BDH(", "BDP(", "BDS(")
_PLUGIN_PREFIX_RE = re.compile(r"\bVA_[A-Z]")

_FUNC_NAME_RE = re.compile(r"([A-Za-z_][A-Za-z0-9_.]*)\s*\(")


class _FuncCollector(Visitor):
    """Visitor that collects function names from a parse tree."""

    def __init__(self):
        self.functions: list[str] = []

    def func_call(self, tree: Tree) -> None:
        name = str(tree.children[0]).upper()
        if name not in self.functions:
            self.functions.append(name)


def _extract_function_names_ast(tree: Tree) -> list[str]:
    """Extract function names via AST walk."""
    collector = _FuncCollector()
    collector.visit(tree)
    return collector.functions


def _extract_function_names_regex(body: str) -> list[str]:
    """Extract function names via regex fallback (for unparseable formulas)."""
    seen: list[str] = []
    for m in _FUNC_NAME_RE.finditer(body):
        name = m.group(1).upper()
        if name not in seen:
            seen.append(name)
    return seen


def find_non_ascii_chars(s: str) -> list[tuple[str, str, int]]:
    """Find non-ASCII characters in a string.

    Returns list of (codepoint, char, count) tuples, e.g.
    [("U+2212", "\u2212", 2), ("U+00A0", "\u00a0", 1)].
    """
    counts: dict[str, int] = {}
    for ch in s:
        if ord(ch) > 127:
            counts[ch] = counts.get(ch, 0) + 1
    result = []
    for ch, count in counts.items():
        cp = f"U+{ord(ch):04X}"
        result.append((cp, ch, count))
    return result


_SANITIZE_MAP = str.maketrans({
    "\u2212": "-",   # minus sign
    "\u2013": "-",   # en dash
    "\u2014": "-",   # em dash
    "\u00a0": " ",   # NBSP
    "\u2018": "'",   # left single quote
    "\u2019": "'",   # right single quote
    "\u201c": '"',   # left double quote
    "\u201d": '"',   # right double quote
})


def sanitize_formula_preview(s: str) -> str:
    """Return a sanitized preview of a formula (diagnostic only, never written to workbook)."""
    return s.translate(_SANITIZE_MAP)


def safe_trim(s: str, n: int) -> str:
    """Trim a string to at most *n* characters, appending '...' if truncated."""
    if len(s) <= n:
        return s
    return s[:n] + "..."


def _format_non_ascii(chars: list[tuple[str, str, int]]) -> str:
    """Format non-ASCII char list for trace output."""
    if not chars:
        return "none"
    parts = []
    for cp, ch, count in chars:
        # Use a readable name if possible
        try:
            name = unicodedata.name(ch)
        except ValueError:
            name = ch
        parts.append(f"{cp}({name}) x{count}")
    return ", ".join(parts)


def classify_formula(formula_str: str) -> dict[str, Any]:
    """Classify a formula string into one of 5 categories.

    Returns dict with:
        classification: "supported" | "unsupported_function" | "parse_error"
                       | "external_link" | "plugin_formula"
        functions_used: list of function names found
        unsupported_functions: list of unsupported function names
        error_message: error string (for parse_error)
    """
    body = formula_str[1:] if formula_str.startswith("=") else formula_str

    result: dict[str, Any] = {
        "classification": "supported",
        "functions_used": [],
        "unsupported_functions": [],
        "error_message": None,
    }

    # 1. Check external link patterns (regex on raw text)
    if _EXTERNAL_LINK_RE.search(body):
        result["classification"] = "external_link"
        result["functions_used"] = _extract_function_names_regex(body)
        return result

    # 2. Check plugin prefixes (BDH/BDP/BDS, VA_)
    upper_body = body.upper()
    if any(prefix in upper_body for prefix in _PLUGIN_PREFIXES) or _PLUGIN_PREFIX_RE.search(body):
        result["classification"] = "plugin_formula"
        result["functions_used"] = _extract_function_names_regex(body)
        return result

    # 3. Attempt parse
    try:
        tree = parse_formula(formula_str)
        funcs = _extract_function_names_ast(tree)
        result["functions_used"] = funcs
        unsupported = [f for f in funcs if f not in _SUPPORTED_FUNCTIONS]
        if unsupported:
            result["classification"] = "unsupported_function"
            result["unsupported_functions"] = unsupported
        else:
            result["classification"] = "supported"
    except Exception as exc:
        result["classification"] = "parse_error"
        result["error_message"] = str(exc)
        result["functions_used"] = _extract_function_names_regex(body)

    return result


def _translate_formula(formula: str) -> str:
    """Best-effort pass-through of an Excel formula.

    We keep the formula as-is since fin123's parser handles common
    arithmetic/function syntax that overlaps with Excel.  No sheet
    cross-references are translated.
    """
    return "=" + formula


def _color_to_hex(color) -> str | None:
    """Extract hex color from an openpyxl Color object.

    Returns a 7-char hex string like ``#ff0000`` or None.
    """
    if color is None:
        return None
    # openpyxl stores theme/indexed/rgb colors
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb)
        # openpyxl may give ARGB (8 chars) or RGB (6 chars)
        if len(rgb) == 8:
            # Skip alpha
            hex_str = "#" + rgb[2:].lower()
        elif len(rgb) == 6:
            hex_str = "#" + rgb.lower()
        else:
            return None
        # Skip black (default) — only return non-default colors
        if hex_str == "#000000":
            return None
        return hex_str
    return None


def _write_trace_log(
    reports_dir: Path,
    all_classifications: list[dict[str, Any]],
) -> None:
    """Write import_trace.log with detailed diagnostics for non-supported formulas."""
    lines: list[str] = ["# fin123 import trace log", ""]
    issues = [c for c in all_classifications if c["classification"] != "supported"]
    if not issues:
        lines.append("No issues found — all formulas classified as supported.")
    else:
        lines.append(f"{len(issues)} issue(s) found:")
        lines.append("")
        for entry in issues:
            cls = entry["classification"]
            sheet = entry["sheet"]
            addr = entry["addr"]
            formula = entry["formula"]
            lines.append(f"[IMPORT][{cls}] {sheet}!{addr}")
            lines.append(f"  formula: {safe_trim(formula, 180)}")
            lines.append(f"  repr   : {entry.get('repr', repr(formula))}")
            lines.append(f"  non_ascii_chars: {entry.get('non_ascii_chars', 'none')}")
            lines.append(f"  sanitized_preview: {entry.get('sanitized_preview', formula)}")
            if cls == "parse_error" and entry.get("error_message"):
                lines.append(f"  parser_error: {entry['error_message']}")
            if cls == "unsupported_function" and entry.get("unsupported_functions"):
                lines.append(f"  functions_detected: {', '.join(entry['unsupported_functions'])}")
            if cls in ("external_link", "plugin_formula"):
                # Show matched pattern hint
                body = formula[1:] if formula.startswith("=") else formula
                if cls == "external_link":
                    m = _EXTERNAL_LINK_RE.search(body)
                    lines.append(f"  match_reason: external link pattern matched: {m.group() if m else 'unknown'}")
                else:
                    upper_body = body.upper()
                    matched = [p for p in _PLUGIN_PREFIXES if p in upper_body]
                    if matched:
                        lines.append(f"  match_reason: plugin prefix: {', '.join(matched)}")
                    elif _PLUGIN_PREFIX_RE.search(body):
                        m = _PLUGIN_PREFIX_RE.search(body)
                        lines.append(f"  match_reason: vendor function pattern: {m.group() if m else 'VA_*'}")
            lines.append("")
    (reports_dir / "import_trace.log").write_text("\n".join(lines))


def import_xlsx(
    xlsx_path: Path,
    target_dir: Path,
    *,
    max_rows: int | None = None,
    max_cols: int | None = None,
    max_total_cells: int | None = None,
) -> dict[str, Any]:
    """Import an XLSX file into a fin123 project directory.

    Creates ``workbook.yaml`` (with sheets) and ``import_report.json``
    inside *target_dir*.

    Args:
        xlsx_path: Path to the .xlsx file.
        target_dir: Destination project directory (created if needed).
        max_rows: Maximum rows per sheet to import (default from config).
        max_cols: Maximum columns per sheet to import (default from config).
        max_total_cells: Maximum total cells across all sheets (default from config).

    Returns:
        The import report dict.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX import.  "
            "Install with: pip install 'fin123[xlsx]'"
        )

    # Load config defaults for limits
    from fin123.project import load_project_config
    if target_dir.exists():
        config = load_project_config(target_dir)
    else:
        from fin123.project import DEFAULT_CONFIG
        config = dict(DEFAULT_CONFIG)

    if max_rows is None:
        max_rows = config.get("max_import_rows_per_sheet", 500)
    if max_cols is None:
        max_cols = config.get("max_import_cols_per_sheet", 100)
    if max_total_cells is None:
        max_total_cells = config.get("max_import_total_cells", 500_000)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)

    sheets: list[dict[str, Any]] = []
    report: dict[str, Any] = {
        "source": str(xlsx_path),
        "sheets_imported": [],
        "cells_imported": 0,
        "formulas_imported": 0,
        "colors_imported": 0,
        "skipped_features": [],
        "warnings": [],
    }

    # Detect unsupported features
    if wb.vba_archive:
        report["skipped_features"].append("VBA macros")
    if hasattr(wb, "chartsheets") and wb.chartsheets:
        report["skipped_features"].append("chart sheets")

    # Classification accumulators (Part A)
    all_classifications: list[dict[str, Any]] = []
    total_cls_counts: Counter = Counter()
    unsupported_func_counts: Counter = Counter()
    total_cells_so_far = 0

    for ws in wb.worksheets:
        # Check total cell limit
        if total_cells_so_far >= max_total_cells:
            report["warnings"].append(
                f"Sheet {ws.title!r}: skipped — total cell limit ({max_total_cells}) reached"
            )
            break

        sheet_name = ws.title
        source_rows = ws.max_row or 1
        source_cols = ws.max_column or 1
        n_rows = min(source_rows, max_rows)
        n_cols = min(source_cols, max_cols)

        # Truncation warnings
        if source_rows > max_rows:
            report["warnings"].append(
                f"Sheet {sheet_name!r}: truncated from {source_rows} to {max_rows} rows"
            )
        if source_cols > max_cols:
            report["warnings"].append(
                f"Sheet {sheet_name!r}: truncated from {source_cols} to {max_cols} columns"
            )

        cells: dict[str, Any] = {}
        fmt: dict[str, Any] = {}
        sheet_formulas = 0
        sheet_colors = 0
        sheet_cls_counts: Counter = Counter()

        # Check for unsupported features on this sheet
        if ws.conditional_formatting:
            report["warnings"].append(
                f"Sheet {sheet_name!r}: conditional formatting skipped"
            )
        if hasattr(ws, "data_validations") and ws.data_validations:
            report["warnings"].append(
                f"Sheet {sheet_name!r}: data validations skipped"
            )

        for row_idx in range(1, n_rows + 1):
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Skip empty cells
                if cell.value is None and cell.data_type == "n":
                    continue
                if cell.value is None:
                    continue

                addr = _make_addr(row_idx - 1, col_idx - 1)

                # Formula?
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    formula_text = str(cell.value)
                    if formula_text.startswith("="):
                        cells[addr] = {"formula": formula_text}
                    else:
                        cells[addr] = {"formula": _translate_formula(formula_text)}
                    sheet_formulas += 1

                    # Classify formula (Part A)
                    stored_formula = cells[addr]["formula"]
                    cls = classify_formula(stored_formula)
                    cls_entry = {
                        "sheet": sheet_name,
                        "addr": addr,
                        "formula": stored_formula,
                        **cls,
                    }
                    # Add diagnostics for non-supported formulas (Part C)
                    if cls["classification"] != "supported":
                        non_ascii = find_non_ascii_chars(stored_formula)
                        cls_entry["repr"] = repr(stored_formula)
                        cls_entry["non_ascii_chars"] = _format_non_ascii(non_ascii)
                        cls_entry["sanitized_preview"] = safe_trim(
                            sanitize_formula_preview(stored_formula), 180
                        )
                    all_classifications.append(cls_entry)
                    sheet_cls_counts[cls["classification"]] += 1
                    total_cls_counts[cls["classification"]] += 1
                    for uf in cls.get("unsupported_functions", []):
                        unsupported_func_counts[uf] += 1

                elif isinstance(cell.value, bool):
                    cells[addr] = {"value": cell.value}
                elif isinstance(cell.value, (int, float)):
                    cells[addr] = {"value": cell.value}
                elif isinstance(cell.value, str):
                    if cell.value:
                        cells[addr] = {"value": cell.value}
                else:
                    # datetime, etc. — convert to string
                    cells[addr] = {"value": str(cell.value)}

                # Font color
                if cell.font and cell.font.color:
                    hex_color = _color_to_hex(cell.font.color)
                    if hex_color:
                        fmt[addr] = {"color": hex_color}
                        sheet_colors += 1

        sheet_data: dict[str, Any] = {
            "name": sheet_name,
            "n_rows": max(n_rows, 200),
            "n_cols": max(n_cols, 40),
            "cells": cells,
        }
        if fmt:
            sheet_data["fmt"] = fmt

        sheets.append(sheet_data)
        report["sheets_imported"].append({
            "name": sheet_name,
            "cells": len(cells),
            "formulas": sheet_formulas,
            "colors": sheet_colors,
            "rows_in_source": source_rows,
            "cols_in_source": source_cols,
            "classifications": dict(sheet_cls_counts),
        })
        report["cells_imported"] += len(cells)
        report["formulas_imported"] += sheet_formulas
        report["colors_imported"] += sheet_colors
        total_cells_so_far += len(cells)

    wb.close()

    # Post-loop warnings
    if report["cells_imported"] > 20_000:
        report["warnings"].append(
            f"Large import: {report['cells_imported']} cells imported"
        )
    ext_link_count = total_cls_counts.get("external_link", 0)
    if ext_link_count > 50:
        report["warnings"].append(
            f"Many external references: {ext_link_count} formulas with external links"
        )

    # Classification summary (Part A)
    total_formulas = sum(total_cls_counts.values())
    report["formula_classifications"] = all_classifications
    report["classification_summary"] = {
        "total_formulas": total_formulas,
        "supported": total_cls_counts.get("supported", 0),
        "parse_errors": total_cls_counts.get("parse_error", 0),
        "unsupported_functions": total_cls_counts.get("unsupported_function", 0),
        "external_links": total_cls_counts.get("external_link", 0),
        "plugin_formulas": total_cls_counts.get("plugin_formula", 0),
    }
    # Top unsupported functions sorted by frequency (max 20)
    report["top_unsupported_functions"] = [
        {"name": name, "count": count}
        for name, count in unsupported_func_counts.most_common(20)
    ]

    # Build workbook spec
    spec: dict[str, Any] = {
        "sheets": sheets,
        "params": {},
        "tables": {},
        "plans": [],
        "outputs": [],
    }

    # Preserve model_id from existing project (if any)
    import yaml as _yaml
    existing_wb_path = target_dir / "workbook.yaml"
    if existing_wb_path.exists():
        existing_spec = _yaml.safe_load(existing_wb_path.read_text()) or {}
        if existing_spec.get("model_id"):
            spec["model_id"] = existing_spec["model_id"]

    # Write output
    target_dir.mkdir(parents=True, exist_ok=True)

    import yaml
    workbook_yaml = yaml.dump(spec, default_flow_style=False, sort_keys=False)
    workbook_path = target_dir / "workbook.yaml"
    workbook_path.write_text(workbook_yaml)

    # Write config
    config_path = target_dir / "fin123.yaml"
    if not config_path.exists():
        config_path.write_text(
            yaml.dump({
                "max_runs": 50,
                "max_artifact_versions": 20,
                "max_total_run_bytes": 2_000_000_000,
                "ttl_days": 30,
            }, default_flow_style=False, sort_keys=False)
        )

    # Standard directories
    for d in ["inputs", "runs", "artifacts", "snapshots", "cache"]:
        (target_dir / d).mkdir(exist_ok=True)

    # Snapshot after import (Part E)
    from fin123.project import ensure_model_id
    from fin123.versioning import SnapshotStore

    ensure_model_id(spec, workbook_path)
    # Re-read after ensure_model_id may have written back
    workbook_yaml = workbook_path.read_text()
    store = SnapshotStore(target_dir)
    version = store.save_snapshot(workbook_yaml)

    # Write report — versioned storage (Part E enhanced)
    report_json = json.dumps(report, indent=2)

    # Directory naming: <timestamp>_import_<n> (Part E)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    import_reports_dir = target_dir / "import_reports"
    import_reports_dir.mkdir(parents=True, exist_ok=True)

    # Find next import number
    existing = sorted(
        d.name for d in import_reports_dir.iterdir()
        if d.is_dir() and "_import_" in d.name
    )
    if existing:
        try:
            last_n = int(existing[-1].rsplit("_", 1)[1])
        except (ValueError, IndexError):
            last_n = 0
        import_n = last_n + 1
    else:
        import_n = 1

    dir_name = f"{ts}_import_{import_n}"
    reports_dir = import_reports_dir / dir_name
    reports_dir.mkdir(parents=True, exist_ok=True)
    (reports_dir / "import_report.json").write_text(report_json)

    # Store source filename (Part E)
    (reports_dir / "source_filename.txt").write_text(xlsx_path.name)

    # Generate import_trace.log
    _write_trace_log(reports_dir, all_classifications)

    # Update import_reports/index.json (Part E enhanced)
    index_path = import_reports_dir / "index.json"
    index: list[dict[str, str]] = []
    if index_path.exists():
        try:
            index = json.loads(index_path.read_text())
        except (json.JSONDecodeError, OSError):
            index = []
    index.append({
        "timestamp": ts,
        "path": f"import_reports/{dir_name}/import_report.json",
        "file": xlsx_path.name,
        "model_version_created": version,
    })
    index_path.write_text(json.dumps(index, indent=2))

    # Backward-compatible root copy
    (target_dir / "import_report.json").write_text(report_json)

    return report
